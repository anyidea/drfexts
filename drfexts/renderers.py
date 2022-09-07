import datetime
import functools
import operator
from decimal import Decimal
from io import BytesIO
from itertools import chain
from typing import Any, Optional

import orjson
import unicodecsv as csv
from django.db.models.query import QuerySet
from django.utils.encoding import force_str
from django.utils.functional import Promise
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import status
from rest_framework.renderers import BaseRenderer
from rest_framework.settings import api_settings
from rest_framework.status import is_success

__all__ = ["CustomJSONRenderer", "CustomCSVRenderer", "CustomXLSXRenderer"]


class CustomJSONRenderer(BaseRenderer):
    """
    Renderer which serializes to JSON.
    Uses the Rust-backed orjson library for serialization speed.
    """

    media_type = "application/json"
    html_media_type = "text/html"
    format = "json"
    charset = None

    options = functools.reduce(
        operator.or_,
        api_settings.user_settings.get("ORJSON_RENDERER_OPTIONS", ()),
        orjson.OPT_SERIALIZE_NUMPY | orjson.OPT_PASSTHROUGH_DATETIME,
    )

    @staticmethod
    def default(obj: Any) -> Any:
        """
        When orjson doesn't recognize an object type for serialization it passes
        that object to this function which then converts the object to its
        native Python equivalent.

        :param obj: Object of any type to be converted.
        :return: native python object
        """

        if isinstance(obj, Promise):
            return force_str(obj)
        elif isinstance(obj, datetime.datetime):
            return obj.strftime(api_settings.DATETIME_FORMAT)
        elif isinstance(obj, Decimal):
            if api_settings.COERCE_DECIMAL_TO_STRING:
                return str(obj)
            else:
                return float(obj)
        elif isinstance(obj, QuerySet):
            return tuple(obj)
        elif hasattr(obj, "tolist"):
            return obj.tolist()
        elif hasattr(obj, "__iter__"):
            return list(item for item in obj)

    def render(
        self,
        data: Any,
        media_type: Optional[str] = None,
        renderer_context: Any = None,
    ) -> bytes:
        """
        Serializes Python objects to JSON.

        :param data: The response data, as set by the Response() instantiation.
        :param media_type: If provided, this is the accepted media type, of the
                `Accept` HTTP header.
        :param renderer_context: If provided, this is a dictionary of contextual
                information provided by the view. By default this will include
                the following keys: view, request, response, args, kwargs
        :return: bytes() representation of the data encoded to UTF-8
        """
        if response := renderer_context.get("response"):
            payload = {}
            if hasattr(renderer_context.get("request"), "id"):
                payload["request_id"] = renderer_context["request"].id

            payload["ret"] = response.status_code
            payload["msg"] = "success"

            if data is not None:
                payload["data"] = data

            if not is_success(response.status_code):
                try:
                    payload["msg"] = data["detail"]
                    payload.pop("data", None)
                except KeyError:
                    payload["msg"] = "Invalid input."
                except TypeError:
                    data = data[0]
                    payload["msg"] = data["detail"]
                    payload.pop("data", None)

            response.status_code = (
                status.HTTP_200_OK
            )  # Set all response status to HTTP 200
        elif data is None:
            return b""
        else:
            payload = data

        # If `indent` is provided in the context, then pretty print the result.
        # E.g. If we're being called by RestFramework's BrowsableAPIRenderer.
        options = self.options
        if media_type == self.html_media_type:
            options |= orjson.OPT_INDENT_2

        response._rendered_data = payload  # for loging response use
        serialized: bytes = orjson.dumps(payload, default=self.default, option=options)
        return serialized


class BaseExportRenderer(BaseRenderer):
    def get_value(self, item, key):
        value = item.get(key, "")
        if isinstance(value, (dict, list)):
            return str(value)

        return value

    def tablize(self, data, header=None):
        """
        Convert a list of data into a table.

        If there is a header provided to tablize it will efficiently yield each
        row as needed. If no header is provided, tablize will need to process
        each row in the data in order to construct a complete header. Thus, if
        you have a lot of data and want to stream it, you should probably
        provide a header to the renderer (using the `header` attribute, or via
        the `renderer_context`).
        """
        # Try to pull the header off of the data, if it's not passed in as an
        # argument.
        if not header and hasattr(data, "header"):
            header = data.header

        if data:
            # First, flatten the data (i.e., convert it to a list of
            # dictionaries that are each exactly one level deep).  The key for
            # each item designates the name of the column that the item will
            # fall into.
            data = self.flatten_data(data)
            # Get the set of all unique headers, and sort them (unless already provided).
            if not header:
                # We don't have to materialize the data generator unless we
                # have to build a header.
                first_data = next(data)
                header = list(first_data.keys())
                data = chain([first_data], data)

            # Return your "table", with the headers as the first row.
            yield header
            # Create a row for each dictionary, filling in columns for which the
            # item has no data with None values.
            for item in data:
                yield (self.get_value(item, key) for key in header)
        elif header:
            # If there's no data but a header was supplied, yield the header.
            yield header
        else:
            # Generator will yield nothing if there's no data and no header
            pass

    def flatten_data(self, data, value_mapping=None):
        """
        Convert the given data collection to a list of dictionaries that are
        each exactly one level deep. The key for each value in the dictionaries
        designates the name of the column that the value will fall into.
        """
        for item in data:
            yield dict(item)


class CustomCSVRenderer(BaseExportRenderer):
    """
    Renderer which serializes to CSV
    """

    media_type = "text/csv"
    format = "csv"
    header = None
    writer_opts = None
    data_key = "results"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        """
        Renders serialized *data* into CSV. For a dictionary:
        """
        renderer_context = renderer_context or {}
        if data is None:
            return bytes()

        if isinstance(data, dict):
            try:
                data = data[self.data_key]
            except KeyError:
                data = []

        writer_opts = renderer_context.get("writer_opts", self.writer_opts or {})
        header = renderer_context.get("header", self.header)
        # excel 打开utf-8的文件会乱码，所以改成gbk
        encoding = renderer_context.get("encoding", "gbk")

        table = self.tablize(data, header=header)
        csv_buffer = BytesIO()
        csv_writer = csv.writer(csv_buffer, encoding=encoding, **writer_opts)
        for row in table:
            csv_writer.writerow(row)

        return csv_buffer.getvalue()


class CustomXLSXRenderer(BaseExportRenderer):
    """
    Renderer for Excel spreadsheet open data format (xlsx).
    """

    media_type = "application/xlsx"
    format = "xlsx"
    header = None
    data_key = "results"
    export_style = {
        "header_font": Font(b=True),
        "header_fill": PatternFill("solid", start_color="87CEFA"),
        "header_alignment": Alignment(vertical="center"),
        "header_height": 18,
        "freeze_header": True,
        "freeze_panes": "A2",
    }

    def render(self, data, accepted_media_type=None, renderer_context=None):
        """
        Render `data` into XLSX workbook, returning a workbook.
        """
        if data is None:
            return bytes()

        if isinstance(data, dict):
            try:
                data = data[self.data_key]
            except KeyError:
                data = []

        header = renderer_context.get("header", self.header)
        export_style = renderer_context.get("export_style", self.export_style)

        table = self.tablize(data, header=header)
        excel_buffer = BytesIO()

        workbook = Workbook()
        sheet = workbook.active

        for row in table:
            sheet.append(row)

        for cell in sheet["1:1"]:
            cell.font = export_style["header_font"]
            cell.fill = export_style["header_fill"]
            cell.alignment = export_style["header_alignment"]

        sheet.row_dimensions[1].height = export_style["header_height"]
        sheet.freeze_panes = export_style.get("freeze_panes", True)

        sheet.print_title_rows = "1:1"
        workbook.save(excel_buffer)

        return excel_buffer.getvalue()
